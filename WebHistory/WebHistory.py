import sqlite3
import argparse
import shutil
import datetime
import os
import platform
import os
import urllib
import openpyxl

debug = False

temp = 'C:\\Users\\' + os.getlogin() + '\\Desktop\\WebHistory\\'
ver = '0.1.0'

def date_from_webkit(webkit_timestamp):
    epoch_start = datetime.datetime(1601,1,1)
    delta = datetime.timedelta(microseconds=int(webkit_timestamp))
    return (epoch_start + delta)

def getChromeHistory(u, wb):
    
    dir=temp+'Chrome\\'
    if os.listdir(dir):
        print(os.listdir(dir))
    else:
        print('empty')

    print('Getting Chrome browser history from:', u)

    chromeHistoryFiles = []
    chromeHistoryFiles.append('C:\\Users\\' + u + '\\AppData\\Local\\Google\\Chrome\\User Data\\Default\\History')
    dir = 'C:\\Users\\' + u + '\\AppData\\Local\\Google\\Chrome\\User Data\\'
    list = os.listdir(path='C:\\Users\\' + u + '\\AppData\\Local\\Google\\Chrome\\User Data\\')
    for file in list:
        if file.startswith('Profile '):
            chromeHistoryFiles.append(dir+file)
            print ('Found profile at: '+dir+file)

    i=0
    for profile in chromeHistoryFiles:

        try:
            i+=1
            print(' Copying '+profile+' to '+temp+'\\Chrome\\History_'+str(i))
            shutil.copyfile(profile,temp+'\\Chrome\\History_'+str(i))
        except Exception as e:
            print(e)
            break
    wsC = wb.create_sheet(title='Chrome')
    wsC.append(['Last Access Time','Base URL','Title','Full URL','Times Accessed'])
    for file in os.listdir(path=temp+'\\Chrome'):
        con = sqlite3.Connection(temp+'\\Chrome\\'+file)
        cur = con.cursor()
        data = cur.execute('''SELECT * FROM urls''')
        #print('======')
        print(file)
        #print('------')
        for row in data:
            #print('Last Access time: ' +  str(date_from_webkit(row[5])))
            #print('Base URL: ' + urllib.parse.urlparse(row[1]).netloc)
            #print('Title: ' + row[2])
            #print('Times Accesed: ' + str(row[3]))
            #print()
            wsC.append({'A':date_from_webkit(row[5]), 'B':urllib.parse.urlparse(row[1]).netloc, 'C':row[2], 'D':row[1], 'E':row[3]})
    
def getFirefoxHistory(u, wb):
    dir=temp+'Firefox\\'
    if os.listdir(dir):
        print(os.listdir(dir))
    else:
        print('empty')

    print('Getting Firefox browser history from:', u)

    firefoxHistoryFiles = []
    firefoxHistoryFiles.append('C:\\Users\\' + u + '\\AppData\\Local\\Mozilla\\Firefox\\User Data\\Default\\History')
    dir = 'C:\\Users\\' + u + '\\AppData\\Local\\Mozilla\\Firefox\\User Data\\'
    try: list = os.listdir(path='C:\\Users\\' + u + '\\AppData\\Local\\Mozilla\\Firefox\\User Data\\')
    except FileNotFoundError:
        print('Firefox data not found!')
        return 0
    for file in list:
        if file.startswith('Profile '):
            firefoxHistoryFiles.append(dir+file)
            print ('Found profile at: '+dir+file)

    i=0
    for profile in firefoxHistoryFiles:

        try:
            i+=1
            print(' Copying '+profile+' to '+temp+'\\Firefox\\History_'+str(i))
            shutil.copyfile(profile,temp+'\\Firefox\\History_'+str(i))
        except Exception as e:
            print(e)
            break
    wsF = wb.create_sheet(title='Firefox')
    wsF.append(['Last Access Time','Base URL','Title','Full URL','Times Accessed'])
    for file in os.listdir(path=temp+'\\Firefox'):
        con = sqlite3.Connection(temp+'\\Firefox\\'+file)
        cur = con.cursor()
        data = cur.execute('''SELECT * FROM urls''')
        #print('======')
        print(file)
        #print('------')
        for row in data:
            #print('Last Access time: ' +  str(date_from_webkit(row[5])))
            #print('Base URL: ' + urllib.parse.urlparse(row[1]).netloc)
            #print('Title: ' + row[2])
            #print('Times Accesed: ' + str(row[3]))
            #print()
            wsF.append({'A':date_from_webkit(row[5]), 'B':urllib.parse.urlparse(row[1]).netloc, 'C':row[2], 'D':row[1], 'E':row[3]})
        
def getEdgeHistory(u, wb):
    dir=temp+'Edge\\'
    if os.listdir(dir):
        print(os.listdir(dir))
    else:
        print('empty')

    print('Getting Edge browser history from:', u)

    edgeHistoryFiles = []
    edgeHistoryFiles.append('C:\\Users\\' + u + '\\AppData\\Local\\Microsoft\\Edge\\User Data\\Default\\History')
    dir = 'C:\\Users\\' + u + '\\AppData\\Local\\Microsoft\\Edge\\User Data\\'
    list = os.listdir(path='C:\\Users\\' + u + '\\AppData\\Local\\Microsoft\\Edge\\User Data\\')
    for file in list:
        if file.startswith('Profile '):
            edgeHistoryFiles.append(dir+file)
            print ('Found profile at: '+dir+file)

    i=0
    for profile in edgeHistoryFiles:

        try:
            i+=1
            print(' Copying '+profile+' to '+temp+'\\Edge\\History_'+str(i))
            shutil.copyfile(profile,temp+'\\Edge\\History_'+str(i))
        except Exception as e:
            print(e)
            break
    wsE = wb.create_sheet(title='Edge')
    wsE.append(['Last Access Time','Base URL','Title','Full URL','Times Accessed'])
    for file in os.listdir(path=temp+'\\Edge'):
        con = sqlite3.Connection(temp+'\\Edge\\'+file)
        cur = con.cursor()
        data = cur.execute('''SELECT * FROM urls''')
        #print('======')
        #print(file)
        #print('------')
        for row in data:
            #print('Last Access time: ' +  str(date_from_webkit(row[5])))
            #print('Base URL: ' + urllib.parse.urlparse(row[1]).netloc)
            #print('Title: ' + row[2])
            #print('Times Accesed: ' + str(row[3]))
            #print()
            wsE.append({'A':date_from_webkit(row[5]), 'B':urllib.parse.urlparse(row[1]).netloc, 'C':row[2], 'D':row[1], 'E':row[3]})

DELETE_ERROR_MSG = 'Cannot delete file. Make sure your have enough credentials to delete this file or that no other process is using this file.'

def main():
    if platform.system() != "Windows":
        print(platform.system())
        raise Exception()

    currentUser = os.getlogin()
    temp = 'C:\\Users\\' + currentUser + '\\Desktop\\WebHistory\\'

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Test"
    
    #command line parser
    parser = argparse.ArgumentParser(description='Doin\' a thing.', prog='WebHistory')
    parser.add_argument('-u', '--username', type=str, required=True, help='doin another thing')
    parser.add_argument('--browser', '-b', choices=['Edge', 'Chrome', 'Firefox', 'All'], default='All')
    parser.add_argument('--version','-v', action='version',version='%(prog)s '+ ver, help='Displays software version')
    
    try:
        if debug:
            args = parser.parse_args(['-v'])
        else:
            args = parser.parse_args()
    except Exception as e:
        #args = parser.parse_args('-h')
        print(e)
        return 0
    os.makedirs(temp, exist_ok=True)

    try: 
        shutil.rmtree(temp+'Chrome')
    except FileNotFoundError:
        print('WebHistory\Chrome pre-cleared!')
    except PermissionError as e:
        print(e)
        return 0
    try: 
        shutil.rmtree(temp+'Firefox')
    except FileNotFoundError:
        print('WebHistory\Firefox pre-cleared!')
    except PermissionError as e:
        print(e)
        return 0
    try: 
        shutil.rmtree(temp+'Edge')
    except FileNotFoundError:
        print('WebHistory\Edge pre-deleted!')
    except PermissionError as e:
        print(e)
        return 0

    #main decision tree
    if args.browser: 
        if args.browser == 'Chrome' or args.browser == 'All':
            try: 
                print('Checking for leftover files from previous run')
                shutil.rmtree(temp+'Chrome')
                print('removed: Webhistory\\Chrome')
            except FileNotFoundError:
                print('Chrome directroy empty - OK')
            except PermissionError:
                print(DELETE_ERROR_MSG)
                return 0
     
            os.makedirs(temp+'Chrome')
            getChromeHistory(args.username, wb)

        if args.browser == 'Firefox' or args.browser == 'All':
            try: 
                print('Checking for leftover files from previous run')
                shutil.rmtree(temp+'Firefox')
                print('removed: Webhistory\\Firefox')
            except FileNotFoundError:
                print('Webhistory\\Firefox OK')
            except PermissionError:
                print(DELETE_ERROR_MSG)
                return 0
     
            os.makedirs(temp+'Firefox')
            getFirefoxHistory(args.username, wb)

        if args.browser == 'Edge' or args.browser == 'All':
            try: 
                print('Checking for leftover files from previous run')
                shutil.rmtree(temp+'Edge')
                print('removed: Webhistory\\Edge')
            except FileNotFoundError:
                print('Webhistory\\Edge OK')
            except PermissionError:
                print(DELETE_ERROR_MSG)
                return 0
     
            os.makedirs(temp+'Edge')
            getEdgeHistory(args.username, wb)


        wb.save(filename = temp+'WebHistory.xlsx')
        print()
        print('Complete!')
        print('See ' +  temp + 'WebHistory.xlsx for data')

    
if __name__ == "__main__":
    main()
